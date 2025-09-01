"""
Excel Report Generator - Fixed version
"""
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
import os, sys

def auto_size(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                v = str(cell.value) if cell.value is not None else ""
                if len(v) > max_length:
                    max_length = len(v)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

def header_style(ws, row=1):
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDDDDD")
    for cell in ws[row]:
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

def write_df(ws, df, start_row=1, start_col=1):
    for j, col in enumerate(df.columns, start_col):
        ws.cell(row=start_row, column=j, value=col)
    header_style(ws, start_row)
    for i, (_, r) in enumerate(df.iterrows(), start_row + 1):
        for j, col in enumerate(df.columns, start_col):
            ws.cell(row=i, column=j, value=r[col])

def build_report(input_csv, output_xlsx, charts_dir=None):
    df = pd.read_csv(input_csv, parse_dates=["Date"])
    df["Month"] = df["Date"].dt.to_period("M").dt.to_timestamp()

    numeric = ["Quantity", "UnitPrice", "Sales"]
    summary = df[numeric].describe().round(2)
    totals = {"Total Sales": float(df["Sales"].sum()), "Total Rows": int(len(df))}

    pivot_region = pd.pivot_table(df, values="Sales", index="Region", aggfunc="sum").sort_values("Sales", ascending=False).round(2)
    pivot_product = pd.pivot_table(df, values="Sales", index=["Category", "Product"], aggfunc="sum").sort_values("Sales", ascending=False).round(2)
    pivot_month = pd.pivot_table(df, values="Sales", index="Month", aggfunc="sum").sort_index().round(2)

    region_png = product_png = month_png = None
    if charts_dir:
        os.makedirs(charts_dir, exist_ok=True)
        # Sales by Region
        plt.figure(figsize=(6,4))
        pivot_region.plot(kind="bar", legend=False)
        plt.title("Sales by Region"); plt.ylabel("Sales"); plt.xlabel("Region"); plt.tight_layout()
        region_png = os.path.join(charts_dir, "sales_by_region.png"); plt.savefig(region_png); plt.close()

        # Top products
        plt.figure(figsize=(8,4))
        top_products = pivot_product.reset_index().head(10)
        labels = top_products["Product"] + " (" + top_products["Category"] + ")"
        plt.bar(labels, top_products["Sales"])
        plt.title("Top Products (by Sales)"); plt.xticks(rotation=45, ha="right"); plt.tight_layout()
        product_png = os.path.join(charts_dir, "top_products.png"); plt.savefig(product_png); plt.close()

        # Monthly sales
        plt.figure(figsize=(8,4))
        pivot_month.plot(kind="line", legend=False, marker='o')
        plt.title("Monthly Sales Trend"); plt.ylabel("Sales"); plt.tight_layout()
        month_png = os.path.join(charts_dir, "monthly_sales.png"); plt.savefig(month_png); plt.close()

    # Build Excel
    wb = Workbook()
    ws_summary = wb.active; ws_summary.title = "Executive Summary"
    ws_region = wb.create_sheet("Pivot_Region")
    ws_product = wb.create_sheet("Pivot_Product")
    ws_month = wb.create_sheet("Pivot_Month")
    ws_data = wb.create_sheet("Raw_Data")

    # Executive summary
    ws_summary["A1"] = "Excel Report"; ws_summary["A1"].font = Font(size=14, bold=True)
    ws_summary["A2"] = "Totals"; ws_summary["A2"].font = Font(bold=True)
    ws_summary["A3"] = "Total Sales"; ws_summary["B3"] = totals["Total Sales"]
    ws_summary["A4"] = "Total Rows"; ws_summary["B4"] = totals["Total Rows"]

    ws_summary["A6"] = "Summary Statistics"; ws_summary["A6"].font = Font(bold=True)
    # write summary dataframe
    r = 7
    for row in dataframe_to_rows(summary.reset_index(), index=False, header=True):
        for c, val in enumerate(row, start=1):
            ws_summary.cell(row=r, column=c, value=val)
        r += 1

    # Insert chart images
    if charts_dir:
        try:
            if region_png and os.path.exists(region_png):
                img = XLImage(region_png); img.width *= 0.7; img.height *= 0.7; ws_summary.add_image(img, "D3")
        except Exception:
            pass
        try:
            if product_png and os.path.exists(product_png):
                img = XLImage(product_png); img.width *= 0.7; img.height *= 0.7; ws_summary.add_image(img, "D20")
        except Exception:
            pass
        try:
            if month_png and os.path.exists(month_png):
                img = XLImage(month_png); img.width *= 0.8; img.height *= 0.8; ws_summary.add_image(img, "D37")
        except Exception:
            pass

    auto_size(ws_summary)

    # Write pivots
    write_df(ws_region, pivot_region.reset_index(), start_row=1, start_col=1); auto_size(ws_region)
    write_df(ws_product, pivot_product.reset_index(), start_row=1, start_col=1); auto_size(ws_product)

    pm = pivot_month.reset_index(); pm["Month"] = pm["Month"].dt.strftime("%Y-%m")
    write_df(ws_month, pm, start_row=1, start_col=1); auto_size(ws_month)

    # Raw data using dataframe_to_rows
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)
    auto_size(ws_data)

    Path(output_xlsx).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    return output_xlsx

if __name__ == "__main__":
    inp = sys.argv[1] if len(sys.argv) > 1 else "sample_sales.csv"
    out = sys.argv[2] if len(sys.argv) > 2 else "output_report.xlsx"
    build_report(inp, out, charts_dir="screenshots")
